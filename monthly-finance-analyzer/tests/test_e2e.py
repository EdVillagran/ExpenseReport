"""End-to-end test: runs the full pipeline on a sample workbook and verifies
that the outputs are correct.

The sample workbook includes:
- Payroll income
- Grocery expense
- Restaurant expense
- Subscription
- Refund (Offset)
- Checking-to-savings transfer (both sides pre-labelled)
- Credit card payment (both sides pre-labelled)
- Late fee
- Unknown merchant (Needs Review)
- Duplicate-looking transaction (same merchant/amount)
"""
from __future__ import annotations

import math
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from finance_analyzer.main import run_pipeline

SAMPLE_WORKBOOK = Path("data/input/sample_e2e.xlsx")
MONTH = "2026-05"


@pytest.fixture()
def pipeline_result(tmp_path):
    result = run_pipeline(SAMPLE_WORKBOOK, MONTH, tmp_path)
    return result


def test_output_workbook_created(pipeline_result, tmp_path):
    assert pipeline_result["workbook_path"].exists(), "Output workbook must be created"


def test_output_report_created(pipeline_result):
    assert pipeline_result["report_path"].exists(), "Markdown report must be created"


def test_output_workbook_has_expected_sheets(pipeline_result):
    wb = openpyxl.load_workbook(pipeline_result["workbook_path"])
    expected_sheets = {
        "Master Transactions",
        "Gross Spending by Category",
        "Necessary vs Unnecessary",
        "Flagged Transactions",
        "Transfers and Payments",
        "Subscriptions",
        "Monthly Summary",
        "Data Quality Issues",
    }
    assert expected_sheets.issubset(set(wb.sheetnames)), (
        f"Missing sheets: {expected_sheets - set(wb.sheetnames)}"
    )


def test_net_spending_is_correct(pipeline_result):
    """Net spending = gross spending - refund offsets.

    Included expenses in sample:
      - Grocery row 1:  150.00
      - Restaurant:      25.00
      - Netflix:         15.99
      - Late fee:        35.00
      - Grocery row 2:  150.00  (duplicate-looking; dates 19 days apart → not auto-excluded)
      Total gross:      375.99
    Refund offset = 30.0
    Net spending = 375.99 - 30.0 = 345.99
    """
    summary = pipeline_result["analytics"]["summary"]
    assert math.isclose(summary["gross_spending"], 375.99, abs_tol=0.01), (
        f"gross_spending expected ~375.99, got {summary['gross_spending']}"
    )
    assert math.isclose(summary["refund_offsets"], 30.0, abs_tol=0.01), (
        f"refund_offsets expected 30.0, got {summary['refund_offsets']}"
    )
    assert math.isclose(summary["net_spending"], 345.99, abs_tol=0.01), (
        f"net_spending expected ~345.99, got {summary['net_spending']}"
    )


def test_transfer_exclusion_is_correct(pipeline_result):
    """Transfer exclusion should be 500 (outgoing side only)."""
    summary = pipeline_result["analytics"]["summary"]
    assert math.isclose(summary["transfers_excluded"], 500.0, abs_tol=0.01), (
        f"transfers_excluded expected 500.0, got {summary['transfers_excluded']}"
    )


def test_credit_card_payment_exclusion_is_correct(pipeline_result):
    """Credit card payment exclusion should be 200 (outgoing side only)."""
    summary = pipeline_result["analytics"]["summary"]
    assert math.isclose(summary["credit_card_payments_excluded"], 200.0, abs_tol=0.01), (
        f"credit_card_payments_excluded expected 200.0, got {summary['credit_card_payments_excluded']}"
    )


def test_review_needed_rows_exist(pipeline_result):
    """There must be at least one Review Needed row (unknown merchant)."""
    summary = pipeline_result["analytics"]["summary"]
    assert summary["review_needed_count"] > 0, "review_needed_count should be > 0"


def test_data_quality_issues_written(pipeline_result):
    """Data quality issues sheet must have at least one issue row."""
    wb = openpyxl.load_workbook(pipeline_result["workbook_path"])
    ws = wb["Data Quality Issues"]
    rows = list(ws.iter_rows(values_only=True))
    # First row is header; there should be at least one data row
    assert len(rows) > 1, "Data Quality Issues sheet should have at least one issue"
    # The single 'no issues' sentinel must NOT appear when there are real issues
    data_values = [str(r[0]) for r in rows[1:]]
    assert "No data quality issues detected" not in data_values
